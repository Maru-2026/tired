from openpyxl import load_workbook
import os

# Directory containing Excel files
folder = r"c:\Users\Maru\OneDrive\Desktop\PT102\New folder"
html_folder = os.path.join(folder, "INDEX.HTML")

# Create HTML content
html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Tables</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #333;
        }
        h2 {
            color: #666;
            margin-top: 30px;
            border-bottom: 2px solid #0066cc;
            padding-bottom: 10px;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 20px 0;
            background-color: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        th {
            background-color: #0066cc;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f0f0f0;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Excel Data Tables</h1>
"""

# Process each Excel file
excel_files = ["Book1.xlsx", "BOOK 2.xlsx", "BOOK 3.xlsx", "BOOK 4.xlsx"]

for excel_file in excel_files:
    filepath = os.path.join(folder, excel_file)
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            sheet_names = wb.sheetnames
            
            html_content += f"\n        <h2>{excel_file}</h2>\n"
            
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                
                rows = list(ws.iter_rows(values_only=True))
                
                if rows:
                    if len(sheet_names) > 1:
                        html_content += f"\n        <h3>Sheet: {sheet_name}</h3>\n"
                    
                    html_content += "        <table>\n"
                    
                    # Add header row
                    if rows:
                        html_content += "            <tr>\n"
                        for cell in rows[0]:
                            html_content += f"                <th>{cell if cell is not None else ''}</th>\n"
                        html_content += "            </tr>\n"
                    
                    # Add data rows
                    for row in rows[1:]:
                        html_content += "            <tr>\n"
                        for cell in row:
                            html_content += f"                <td>{cell if cell is not None else ''}</td>\n"
                        html_content += "            </tr>\n"
                    
                    html_content += "        </table>\n"
        except Exception as e:
            html_content += f"<p style='color: red;'>Error reading {excel_file}: {str(e)}</p>\n"

# Close HTML
html_content += """
    </div>
</body>
</html>
"""

# Write to INDEX.HTML folder
output_file = os.path.join(html_folder, "index.html")
os.makedirs(html_folder, exist_ok=True)
with open(output_file, "w", encoding="utf-8") as f:
    f.write(html_content)

print(f"HTML file created successfully at: {output_file}")
